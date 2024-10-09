# -*- coding:utf-8 -*-
# kmeans : Bisecting k-means cluster(二分K-means算法)
from openpyxl import load_workbook
import openpyxl
import math
import random
from sklearn import datasets
import numpy as np
import pandas as pd
from numpy import *
from matplotlib import pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from sklearn.cluster import KMeans

plt.rcParams['font.family'] = ['sans-serif']
plt.rcParams['font.sans-serif'] = ['SimHei']

# 打开Excel文件
workbook = load_workbook(filename="12、17、21年总得分.xlsx")
# 获取活动工作表
worksheet = workbook.active
data = []
for column in range(2, 32):
    one_data = worksheet.cell(row=2, column=column).value
    data.append(one_data)

data = np.array(data)
data = data.reshape(-1, 1)

# 生成城市列表
cities = ['Beijing', 'Tianjin', 'Hebei', 'Shanxi', 'Neimenggu', 'Liaonin', 'Jilin', 'Heilongjiang', 'Shanghai', 'Jiangsu', 'Zhejiang',
          'Anhui', 'Fujian', 'Jiangxi', 'Shandong', 'Henan', 'Hubei', 'Hunan', 'Guangdong', 'Guangxi', 'Hainan', 'Chongqin', 'Sichuan', 'Guizhou',
    'Yunnan', 'Shanxi', 'Gansu', 'Qinghai', 'Ningxia', 'Xinjiang']

# 使用K-Means算法进行聚类
kmeans = KMeans(n_clusters=3, init='k-means++', max_iter=300, n_init=10, random_state=0)
kmeans.fit(data)

# 绘制图形
cities = list(range(30))
colors = ['red', 'green', 'blue']
markers = ['x', 'x', 'x']  # 第一类、第二类和第三类的质心使用叉号标记


fig, ax = plt.subplots(figsize=(8, 6))

for i in range(3):
    x = []
    y = []
    for j in range(30):
        if kmeans.labels_[j] == i:
            x.append(cities[j])
            y.append(data[j])
            z = j
    ax.scatter(x, y, c=colors[i], marker='o')
    centroid = kmeans.cluster_centers_[i]
    ax.scatter(z, centroid[0], c=colors[i], marker=markers[i], s=100)

# 添加标题和标签
ax.set_title('2012_Result', fontsize=20)
ax.set_xlabel('Cities')
ax.set_ylabel('Scores')

colors = ['r', 'g', 'b']
labels = ['第一组', '第二组', '第三组']
handles = [plt.Line2D([], [], color=c, label=l) for c, l in zip(colors, labels)]
ax.legend(handles=handles)

# 添加注释
fig.text(0.98, 0.02, '编号对应城市', transform=ax.transAxes, ha='right', va='bottom', fontsize=10)

# 显示图形
plt.show()


for i in range(3):
    centroid = kmeans.cluster_centers_[i]
    print(f"Cluster {i+1} centroid: {centroid[0]}")

df = pd.DataFrame(data, columns=['Score'])
df['Cluster'] = kmeans.labels_
print(df)