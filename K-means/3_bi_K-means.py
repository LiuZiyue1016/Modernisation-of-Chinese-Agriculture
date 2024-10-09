from numpy import *
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import openpyxl
import statistics
import pandas as pd
import numpy as np

def load_data_set():
    """加载数据集"""
    # 打开Excel文件
    workbook = load_workbook(filename="12、17、21年总得分_1.xlsx")

    # 获取活动工作表
    worksheet = workbook.active
    dataSet = []
    for row in range(2, 32):
        one_data = [worksheet.cell(row=row, column=2).value, worksheet.cell(row=row, column=3).value,
                    worksheet.cell(row=row, column=4).value]
        dataSet.append(one_data)

    return dataSet

def distance_euclidean(vector1, vector2):
    """计算欧氏距离"""
    return sqrt(sum(power(vector1-vector2, 2)))  # 返回两个向量的距离

def rand_center(dataSet, k):
    """构建一个包含K个随机质心的集合"""
    n = shape(dataSet)[1]  # 获取样本特征值
    # 初始化质心，创建(k,n)个以0填充的矩阵
    centroids = mat(zeros((k, n)))  # 每个质心有n个坐标值，总共要k个质心
    # 遍历特征值
    for j in range(n):
        # 计算每一列的最小值
        minJ = min(dataSet[:, j])
        # 计算每一列的范围值
        rangeJ = float(max(dataSet[:, j]) - minJ)
        # 计算每一列的质心，并将其赋给centroids
        centroids[:, j] = minJ + rangeJ * random.rand(k, 1)
    return centroids   # 返回质心

def k_means(dataSet, k, distMeas = distance_euclidean, creatCent = rand_center):
    """K-means聚类算法"""
    m = shape(dataSet)[0] # 行数
    # 建立簇分配结果矩阵，第一列存放该数据所属中心点，第二列是该数据到中心点的距离
    clusterAssment = mat(zeros((m, 2)))
    centroids = creatCent(dataSet, k)  # 质心，即聚类点
    # 用来判定聚类是否收敛
    clusterChanged = True
    while clusterChanged:
        clusterChanged = False
        for i in range(m):  # 把每一个数据划分到离他最近的中心点
            minDist = inf  # 无穷大
            minIndex = -1  #初始化
            for j in range(k):
                # 计算各点与新的聚类中心的距离
                distJI = distMeas(centroids[j, :], dataSet[i, :])
                if distJI < minDist:
                    # 如果第i个数据点到第j中心点更近，则将i归属为j
                    minDist = distJI
                    minIndex = j
            # 如果分配发生变化，则需要继续迭代
            if clusterAssment[i, 0] != minIndex:
                clusterChanged = True
            # 并将第i个数据点的分配情况存入字典
            clusterAssment[i, :] = minIndex, minDist**2
        # print(centroids)
        for cent in range(k):  # 重新计算中心点
            # 去第一列等于cent的所有列
            ptsInClust = dataSet[nonzero(clusterAssment[:, 0].A == cent)[0]]
            # 算出这些数据的中心点
            centroids[cent, :] = mean(ptsInClust, axis=0)
    return centroids, clusterAssment

def biKmeans(dataMat, k, distMeas=distance_euclidean):
    """二分k-means算法"""
    m = shape(dataMat)[0]
    # 创建一个矩阵来存储数据集中每个点的簇分配结果及平方误差
    clusterAssment = mat(zeros((m, 2)))
    # 根据数据集均值获取第一个质心
    centroid0 = mean(dataMat, axis=0).tolist()[0]
    # 用一个列表来保留所有的质心
    centList = [centroid0]
    # 遍历数据集中所有点来计算每个点到质心的距离
    for j in range(m):
        clusterAssment[j, 1] = distMeas(mat(centroid0), dataMat[j, :]) ** 2
    # 对簇不停的进行划分,直到得到想要的簇数目为止
    while (len(centList) < k):
        # 初始化最小SSE为无穷大,用于比较划分前后的SSE
        lowestSSE = inf  # 无穷大
        # 通过考察簇列表中的值来获得当前簇的数目,遍历所有的簇来决定最佳的簇进行划分
        for i in range(len(centList)):
            # 对每一个簇,将该簇中的所有点看成一个小的数据集
            ptsInCurrCluster = dataMat[nonzero(clusterAssment[:, 0].A == i)[0], :]
            # 将ptsInCurrCluster输入到函数kMeans中进行处理,k=2,
            # kMeans会生成两个质心(簇),同时给出每个簇的误差值
            centroidMat, splitClustAss = k_means(ptsInCurrCluster, 2, distMeas)
            # 划分数据的SSE与未划分的之和作为本次划分的总误差
            sseSplit = sum(splitClustAss[:, 1])  # 划分数据集的SSE
            sseNotSplit = sum(clusterAssment[nonzero(clusterAssment[:, 0].A != i)[0], 1])  # 未划分数据集的SSE
            print('划分数据集的SSE, and 未划分的SSE: ', sseSplit, sseNotSplit)
            # 将划分与未划分的SSE求和与最小SSE相比较 确定是否划分
            if (sseSplit + sseNotSplit) < lowestSSE:
                bestCentToSplit = i  # 当前最适合做划分的中心点
                bestNewCents = centroidMat  # 划分后的两个新中心点
                bestClustAss = splitClustAss.copy()  # 划分点的聚类信息
                lowestSSE = sseSplit + sseNotSplit
        # 找出最好的簇分配结果
        # 调用kmeans函数并且指定簇数为2时,会得到两个编号分别为0和1的结果簇
        bestClustAss[nonzero(bestClustAss[:, 0].A == 1)[0], 0] = len(centList)
        # 更新为最佳质心
        bestClustAss[nonzero(bestClustAss[:, 0].A == 0)[0], 0] = bestCentToSplit
        print('本次最适合划分的质心: ', bestCentToSplit)
        print('被划分数据集样本数量: ', len(bestClustAss))
        # 更新质心列表
        # 更新原质心list中的第i个质心为使用二分kMeans后bestNewCents的第一个质心
        centList[bestCentToSplit] = bestNewCents[0, :].tolist()[0]
        # 添加bestNewCents的第二个质心
        centList.append(bestNewCents[1, :].tolist()[0])
        # 重新分配最好簇下的数据(质心)以及SSE
        clusterAssment[nonzero(clusterAssment[:, 0].A == bestCentToSplit)[0], :] = bestClustAss
    return mat(centList), clusterAssment

def draw_save_result(datMat, clusterAssment):

    datMat = array(datMat)
    clusterAssment = clusterAssment.astype('U')
    colors = []
    m = shape(datMat)[0]

    for row in range(m):
        if clusterAssment[row, 0] == '0.0':
            colors.append('#FF0000')
        elif clusterAssment[row, 0] == '1.0':
            colors.append('#00FF00')
        else:
            colors.append('#0000FF')
    colors.append('#000000')
    fig = plt.figure()
    ax = fig.add_subplot(111, projection='3d')
    for row in range(30):
        ax.scatter(datMat[row, 0], datMat[row, 1], datMat[row, 2], c=colors[row])
        ax.scatter(centList[:, 0].tolist(), centList[:, 1].tolist(), centList[:, 2].tolist(), c=colors[30], marker='x')

    ax.set_xlabel('the12th year')
    ax.set_ylabel('the17th year')
    ax.set_zlabel('the21th year')

    plt.title("result")
    plt.savefig('result')
    plt.show()

def find_city(datMat, clusterAssment, cenlist, distMeas=distance_euclidean):
    m = shape(datMat)[0]
    clusterAssment = clusterAssment.astype('U')
    distance_0 = []
    distance_1 = []
    distance_2 = []
    for row in range(m):
        if clusterAssment[row, 0] == '0.0':
            distance_0_point = distMeas(datMat[row, :], centList[0]) ** 2
            distance_0.append(distance_0_point)
            if distance_0_point == min(distance_0):
                city_0 = row

        elif clusterAssment[row, 0] == '1.0':
            distance_1_point = distMeas(datMat[row, :], centList[0]) ** 2
            distance_1.append(distance_1_point)
            if distance_1_point == min(distance_1):
                city_1 = row

        elif clusterAssment[row, 0] == '2.0':
            distance_2_point = distMeas(datMat[row, :], centList[2]) ** 2
            distance_2.append(distance_2_point)
            if distance_2_point == min(distance_2):
                city_2 = row

    citys = [city_0, city_1, city_2]
    return citys


datMat = mat(load_data_set())
centList, clusterAssment = biKmeans(datMat, 3)
print("质心结果：", centList)
print("聚类结果：", clusterAssment)
citys = find_city(datMat, clusterAssment, centList)
print("最接近质点的城市：", citys)
draw_save_result(datMat, clusterAssment)

