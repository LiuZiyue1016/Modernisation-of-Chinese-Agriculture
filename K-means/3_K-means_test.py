import random
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import openpyxl
import statistics
import pandas as pd
import imageio
from PIL import Image


" 计算质心到点的距离"
def calcDis(dataSet, centroids, k):
    clalist=[]
    for data in dataSet:
        diff = np.tile(data, (k, 1)) - centroids  #相减   (np.tile(a,(2,1))就是把a先沿x轴复制1倍，即没有复制，仍然是 [0,1,2]。 再把结果沿y方向复制2倍得到array([[0,1,2],[0,1,2]]))
        squaredDiff = diff ** 2     #平方
        squaredDist = np.sum(squaredDiff, axis=1)   #和  (axis=1表示行)
        distance = squaredDist ** 0.5  #开根号
        clalist.append(distance)
    clalist = np.array(clalist)  #返回一个每个点到质点的距离len(dateSet)*k的数组

    return clalist


""" 计算质心，求出样本属于哪一个簇 """
def classify(dataSet, centroids, k):
    # 计算样本到质心的距离
    clalist = calcDis(dataSet, centroids, k)
    # 分组并计算新的质心
    minDistIndices = np.argmin(clalist, axis=1)  # axis=1 表示求出每行的最小值的下标
    newCentroids = pd.DataFrame(dataSet).groupby(
        minDistIndices).mean()  # DataFramte(dataSet)对DataSet分组，groupby(min)按照min进行统计分类，mean()对分类结果求均值
    newCentroids = newCentroids.values
    # 计算变化量
    changed = newCentroids - centroids

    return changed, newCentroids


""" 点的簇分布 --> 展示出来"""


def showCluster(cluster, centroids, k):
    """
    绘制样本点
    """
    # 用不同颜色形状来表示各个类别
    colors = ['#FF0000', '#FFA500', '#FF4500', '#00FF00', '#00FFFF', '#0000FF', '#800080', '#808080', '#A52A2A']
    fig = plt.figure()
    ax = fig.add_subplot(111, projection='3d')

    for i, points in enumerate(cluster):  # 取出簇
        #因为同一个簇的需要绘制同样的颜色，所以需要得到同一个簇的索引
        color = colors[i]
        for point in points:  # 取出属于簇中的每个样本点
            ax.scatter(point[0], point[1], point[2], c=color)

    """
    绘制质心点
    """
    # 用不同颜色形状来表示各个类别
    # 绘制质心点
    for j in range(len(centroids)):
        color = colors[j]
        ax.scatter(centroids[j][0], centroids[j][1], centroids[j][2], c=color, marker='x')

    ax.set_xlabel('the12th year')
    ax.set_ylabel('the17th year')
    ax.set_zlabel('the21th year')

    plt.title(f"The change of k")
    plt.savefig(f'k={k}.png')
    plt.show()


"""封装并使用k-means分类"""


def kmeans(dataSet, k):
    # 随机取质心
    centroids = random.sample(dataSet, k)

    # 更新质心 直到变化量全为0
    changed, newCentroids = classify(dataSet, centroids, k)
    while np.any(changed != 0):
        changed, newCentroids = classify(dataSet, newCentroids, k)

    centroids = sorted(newCentroids.tolist())  # tolist()将矩阵转换成列表 sorted()排序

    # 根据质心计算每个集群
    cluster = []
    clalist = calcDis(dataSet, centroids, k)  # 调用欧拉距离
    minDistIndices = np.argmin(clalist, axis=1)

    for i in range(k):
        cluster.append([])
    for i, j in enumerate(minDistIndices):  # enumerate()可同时遍历索引和遍历元素 i对应索引，j对应元素（0,...,k-1），而元素对应的刚好是k
        cluster[j].append(dataSet[i])

    # 展示点的簇分布
    showCluster(cluster, centroids, k)

    # 计算平均损失，判断不同k的分类效果
    minDistance = np.min(clalist, axis=1)  # axis=1 表示求出每行的最小值
    loss = np.mean(minDistance)  # # 计算距离变化值（损失）  样本跟它所属簇的误差求和/样本个数-->平均损失

    return centroids, cluster, loss

"""创建数据集 """
def createDataSet():
    # 打开Excel文件
    workbook = load_workbook(filename="12、17、21年总得分_1.xlsx")
    # 获取活动工作表
    worksheet = workbook.active
    data = []
    for row in range(2, 32):
        one_data = [worksheet.cell(row=row, column=2).value, worksheet.cell(row=row, column=3).value,
                    worksheet.cell(row=row, column=4).value]
        data.append(one_data)

    # 声明一个 3D 图形对象
    fig = plt.figure()
    ax = fig.add_subplot(111, projection='3d')
    # 将数据放入三个坐标轴中
    x = [data[i][0] for i in range(len(data))]
    y = [data[i][1] for i in range(len(data))]
    z = [data[i][2] for i in range(len(data))]
    # 绘制散点图
    ax.scatter(x, y, z)
    # 设置坐标轴标签
    ax.set_xlabel('the12th year')
    ax.set_ylabel('the17th year')
    ax.set_zlabel('the21th year')

    plt.title("data")
    plt.savefig('data.png')
    # 显示图形
    plt.show()

    data = np.array(data)
    data = data.tolist()
    return data


if __name__ == '__main__':
    dataset = createDataSet()

    # 存储每个k值下的损失
    loss_list = []

    for k in range(2, 10):  # k的范围在2，9
        centroids, cluster, loss = kmeans(dataset, k)
        print('k为：', k)
        print('质心为：%s' % centroids)
        print('集群为：%s' % cluster)
        loss_list.append(loss)

    # 观察k值与损失的关系
    plt.figure(figsize=(8, 6))
    plt.plot(range(2, 10), loss_list)  # 绘制不同k下的效果
    plt.xlabel('k')
    plt.ylabel("loss")
    plt.title("loss/k")
    plt.savefig('loss-k.png')
    plt.show()

    images = []
    # 将每一帧图像写入到动画中
    for i in range(2, 10):
        filename = f'k={i}.png'
        images.append(imageio.imread(filename))

    imageio.mimsave('change.gif', images, duration=600)